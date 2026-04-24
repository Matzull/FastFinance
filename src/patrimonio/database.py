"""Database manager for the personal finance application."""

import csv
import io
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from patrimonio.models import (
    Banco,
    Base,
    Frecuencia,
    Patrimonio,
    PeriodoPresupuesto,
    Presupuesto,
    Suscripcion,
    TipoTransaccion,
    Transaccion,
)

COLUMN_ALIASES = {
    "date": {
        "date",
        "bookingdate",
        "valuedate",
        "transactiondate",
        "operationdate",
        "fecha",
        "fechacontable",
        "fechaoperacion",
        "fechavalor",
        "dia",
    },
    "description": {
        "description",
        "details",
        "detail",
        "concept",
        "concepto",
        "merchant",
        "comercio",
        "movimiento",
        "descripcion",
        "descripción",
        "glosa",
        "narration",
    },
    "amount": {
        "amount",
        "importe",
        "cantidad",
        "monto",
        "valor",
        "total",
        "importeeur",
    },
    "debit": {
        "debit",
        "charge",
        "withdrawal",
        "outflow",
        "cargo",
        "debe",
        "debito",
        "débito",
    },
    "credit": {
        "credit",
        "deposit",
        "inflow",
        "abono",
        "haber",
        "credito",
        "crédito",
    },
    "category": {
        "category",
        "categoria",
        "categoría",
        "tag",
        "etiqueta",
    },
    "direction": {
        "type",
        "direction",
        "movement",
        "kind",
        "nature",
        "tipo",
        "naturaleza",
        "sentido",
    },
}

EXPENSE_DIRECTION_WORDS = {
    "expense",
    "out",
    "outgoing",
    "debit",
    "charge",
    "payment",
    "withdrawal",
    "gasto",
    "cargo",
    "debito",
    "débito",
    "salida",
    "pago",
}


def _normalize_header(header: str) -> str:
    ascii_header = (
        unicodedata.normalize("NFKD", str(header)).encode("ascii", "ignore").decode("ascii")
    )
    return "".join(ch for ch in ascii_header.lower() if ch.isalnum())


def _looks_like_expense(direction_value: str | None) -> bool:
    if direction_value is None:
        return False
    normalized = _normalize_header(direction_value)
    return normalized in {_normalize_header(word) for word in EXPENSE_DIRECTION_WORDS}


class GestorDB:
    """Handles all database operations."""

    def __init__(self, db_path: Path | None = None):
        if db_path is None:
            db_path = Path.home() / ".patrimonio" / "patrimonio.db"
        db_path.parent.mkdir(parents=True, exist_ok=True)

        self.engine = create_engine(f"sqlite:///{db_path}", echo=False)
        Base.metadata.create_all(self.engine)

    def get_session(self) -> Session:
        return Session(self.engine)

    def _resolve_column_indexes(self, headers: list[str]) -> dict[str, int]:
        normalized_headers = [_normalize_header(header) for header in headers]
        indexes: dict[str, int] = {}

        for key, aliases in COLUMN_ALIASES.items():
            normalized_aliases = {_normalize_header(alias) for alias in aliases}
            for index, header in enumerate(normalized_headers):
                if header in normalized_aliases:
                    indexes[key] = index
                    break

        if "date" not in indexes or "description" not in indexes:
            raise ValueError("Could not detect required columns: date and description")
        if "amount" not in indexes and "debit" not in indexes:
            raise ValueError("Could not detect amount column. Expected amount or debit")

        return indexes

    def _parse_decimal(self, value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        text = str(value).strip()
        if not text:
            return None
        cleaned = text.replace("€", "").replace("$", "").replace(" ", "")

        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")

        try:
            return Decimal(cleaned)
        except Exception:
            return None

    def _parse_date(self, value: object) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        text = str(value).strip()
        if not text:
            return None

        formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%d/%m/%y",
            "%d-%m-%y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None

    def _read_tabular_file(
        self, file_name: str, file_bytes: bytes
    ) -> tuple[list[str], list[list[object]]]:
        suffix = Path(file_name).suffix.lower()

        if suffix == ".csv":
            content = file_bytes.decode("utf-8-sig", errors="replace")
            sample = content[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel

            rows = list(csv.reader(io.StringIO(content), dialect=dialect))
            if not rows:
                raise ValueError("CSV file is empty")
            return [str(value).strip() for value in rows[0]], [list(row) for row in rows[1:]]

        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            worksheet = workbook.active
            all_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not all_rows:
                raise ValueError("Excel file is empty")
            headers = [str(value).strip() if value is not None else "" for value in all_rows[0]]
            return headers, all_rows[1:]

        raise ValueError("Unsupported file format. Allowed: .csv, .xlsx")

    def import_statement_expenses(
        self,
        bank_id: int,
        file_name: str,
        file_bytes: bytes,
        default_category: str = "otros",
    ) -> dict:
        """Import expenses from a bank statement file (CSV/XLSX)."""
        if not self.get_bank(bank_id):
            raise ValueError("Bank account not found")

        headers, rows = self._read_tabular_file(file_name=file_name, file_bytes=file_bytes)
        indexes = self._resolve_column_indexes(headers)

        imported_count = 0
        skipped_count = 0
        errors: list[str] = []

        for row_number, row in enumerate(rows, start=2):
            row_values = list(row)
            if not any(value is not None and str(value).strip() for value in row_values):
                continue

            def value_for(column_key: str) -> object | None:
                index = indexes.get(column_key)
                if index is None or index >= len(row_values):
                    return None
                return row_values[index]

            parsed_date = self._parse_date(value_for("date")) or date.today()
            description = str(value_for("description") or "").strip()
            if not description:
                description = "Imported expense"

            debit_amount = self._parse_decimal(value_for("debit"))
            raw_amount = self._parse_decimal(value_for("amount"))
            direction_value = value_for("direction")

            amount: Decimal | None = None
            if debit_amount is not None and debit_amount > 0:
                amount = debit_amount
            elif raw_amount is not None:
                if raw_amount < 0:
                    amount = abs(raw_amount)
                elif raw_amount > 0 and _looks_like_expense(
                    str(direction_value) if direction_value else None
                ):
                    amount = raw_amount

            if amount is None or amount <= 0:
                skipped_count += 1
                continue

            category_value = str(value_for("category") or "").strip().lower() or default_category

            try:
                self.create_transaction(
                    banco_id=bank_id,
                    tipo=TipoTransaccion.GASTO,
                    cantidad=amount,
                    descripcion=description,
                    categoria=category_value,
                    fecha=parsed_date,
                )
                imported_count += 1
            except Exception as exc:
                skipped_count += 1
                errors.append(f"row {row_number}: {exc}")

        return {
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "detected_columns": {key: headers[indexes[key]] for key in indexes},
        }

    def import_expenses_from_statement_compat(
        self,
        banco_id: int,
        file_name: str,
        file_bytes: bytes,
        categoria_default: str = "otros",
    ) -> dict:
        """Compatibility wrapper for previous Spanish argument names."""
        return self.import_statement_expenses(
            bank_id=banco_id,
            file_name=file_name,
            file_bytes=file_bytes,
            default_category=categoria_default,
        )

    # ==================== BANKS ====================

    def create_bank(
        self,
        nombre: str,
        tipo_cuenta: str,
        saldo_inicial: Decimal = Decimal("0"),
        moneda: str = "EUR",
        notas: str | None = None,
    ) -> Banco:
        """Creates a new bank account."""
        with self.get_session() as session:
            banco = Banco(
                nombre=nombre,
                tipo_cuenta=tipo_cuenta,
                saldo_inicial=saldo_inicial,
                moneda=moneda,
                notas=notas,
            )
            session.add(banco)
            session.commit()
            session.refresh(banco)
            return banco

    def list_banks(self, solo_activos: bool = True) -> list[Banco]:
        """Lists all bank accounts."""
        with self.get_session() as session:
            query = select(Banco)
            if solo_activos:
                query = query.where(Banco.activo)
            return list(session.scalars(query).all())

    def get_bank(self, banco_id: int) -> Banco | None:
        """Gets a bank account by ID."""
        with self.get_session() as session:
            return session.get(Banco, banco_id)

    def calculate_bank_balance(self, banco_id: int) -> Decimal:
        """Calculates current balance for a bank account."""
        with self.get_session() as session:
            banco = session.get(Banco, banco_id)
            if not banco:
                return Decimal("0")

            saldo = banco.saldo_inicial

            # Add income
            ingresos = session.scalar(
                select(func.sum(Transaccion.cantidad))
                .where(Transaccion.banco_id == banco_id)
                .where(Transaccion.tipo == TipoTransaccion.INGRESO)
            ) or Decimal("0")

            # Subtract expenses
            gastos = session.scalar(
                select(func.sum(Transaccion.cantidad))
                .where(Transaccion.banco_id == banco_id)
                .where(Transaccion.tipo == TipoTransaccion.GASTO)
            ) or Decimal("0")

            return saldo + ingresos - gastos

    # ==================== TRANSACTIONS ====================

    def create_transaction(
        self,
        banco_id: int,
        tipo: TipoTransaccion,
        cantidad: Decimal,
        descripcion: str,
        categoria: str,
        fecha: date | None = None,
        notas: str | None = None,
    ) -> Transaccion:
        """Creates a new transaction."""
        with self.get_session() as session:
            transaccion = Transaccion(
                banco_id=banco_id,
                tipo=tipo,
                cantidad=cantidad,
                descripcion=descripcion,
                categoria=categoria,
                fecha=fecha or date.today(),
                notas=notas,
            )
            session.add(transaccion)
            session.commit()
            session.refresh(transaccion)
            return transaccion

    def list_transactions(
        self,
        banco_id: int | None = None,
        tipo: TipoTransaccion | None = None,
        categoria: str | None = None,
        fecha_desde: date | None = None,
        fecha_hasta: date | None = None,
        limite: int = 50,
    ) -> list[Transaccion]:
        """Lists transactions with optional filters."""
        with self.get_session() as session:
            query = select(Transaccion).order_by(Transaccion.fecha.desc())

            if banco_id:
                query = query.where(Transaccion.banco_id == banco_id)
            if tipo:
                query = query.where(Transaccion.tipo == tipo)
            if categoria:
                query = query.where(Transaccion.categoria == categoria)
            if fecha_desde:
                query = query.where(Transaccion.fecha >= fecha_desde)
            if fecha_hasta:
                query = query.where(Transaccion.fecha <= fecha_hasta)

            query = query.limit(limite)
            return list(session.scalars(query).all())

    def delete_transaction(self, transaccion_id: int) -> bool:
        """Deletes a transaction."""
        with self.get_session() as session:
            transaccion = session.get(Transaccion, transaccion_id)
            if transaccion:
                session.delete(transaccion)
                session.commit()
                return True
            return False

    # ==================== SUBSCRIPTIONS ====================

    def create_subscription(
        self,
        banco_id: int,
        nombre: str,
        cantidad: Decimal,
        frecuencia: Frecuencia,
        fecha_inicio: date | None = None,
        categoria: str = "suscripciones",
        notas: str | None = None,
    ) -> Suscripcion:
        """Creates a new subscription."""
        with self.get_session() as session:
            suscripcion = Suscripcion(
                banco_id=banco_id,
                nombre=nombre,
                cantidad=cantidad,
                frecuencia=frecuencia,
                fecha_inicio=fecha_inicio or date.today(),
                categoria=categoria,
                notas=notas,
            )
            session.add(suscripcion)
            session.commit()
            session.refresh(suscripcion)
            return suscripcion

    def list_subscriptions(self, solo_activas: bool = True) -> list[Suscripcion]:
        """Lists all subscriptions."""
        with self.get_session() as session:
            query = select(Suscripcion)
            if solo_activas:
                query = query.where(Suscripcion.activa)
            return list(session.scalars(query).all())

    def cancel_subscription(self, suscripcion_id: int) -> bool:
        """Cancels a subscription."""
        with self.get_session() as session:
            suscripcion = session.get(Suscripcion, suscripcion_id)
            if suscripcion:
                suscripcion.activa = False
                suscripcion.fecha_fin = date.today()
                session.commit()
                return True
            return False

    def calculate_monthly_subscription_cost(self) -> Decimal:
        """Calculates total monthly subscription cost."""
        suscripciones = self.list_subscriptions(solo_activas=True)
        total = Decimal("0")
        for sub in suscripciones:
            total += sub.costo_mensual()
        return total

    # ==================== NET WORTH ====================

    def create_net_worth_item(
        self,
        nombre: str,
        tipo: str,
        valor: Decimal,
        descripcion: str | None = None,
        fecha_adquisicion: date | None = None,
    ) -> Patrimonio:
        """Creates a new net worth item."""
        with self.get_session() as session:
            patrimonio = Patrimonio(
                nombre=nombre,
                tipo=tipo,
                valor=valor,
                descripcion=descripcion,
                fecha_adquisicion=fecha_adquisicion,
            )
            session.add(patrimonio)
            session.commit()
            session.refresh(patrimonio)
            return patrimonio

    def list_net_worth_items(self, tipo: str | None = None) -> list[Patrimonio]:
        """Lists all net worth items."""
        with self.get_session() as session:
            query = select(Patrimonio)
            if tipo:
                query = query.where(Patrimonio.tipo == tipo)
            return list(session.scalars(query).all())

    def calculate_net_worth(self) -> Decimal:
        """Calculates net worth (assets minus liabilities)."""
        with self.get_session() as session:
            activos = session.scalar(
                select(func.sum(Patrimonio.valor)).where(Patrimonio.tipo == "activo")
            ) or Decimal("0")

            pasivos = session.scalar(
                select(func.sum(Patrimonio.valor)).where(Patrimonio.tipo == "pasivo")
            ) or Decimal("0")

            # Add bank balances
            for banco in self.list_banks():
                activos += self.calculate_bank_balance(banco.id)

            return activos - pasivos

    # ==================== SUMMARIES ====================

    def monthly_summary(self, mes: int, año: int) -> dict:
        """Generates a summary for the specified month."""
        fecha_inicio = date(año, mes, 1)
        if mes == 12:
            fecha_fin = date(año + 1, 1, 1)
        else:
            fecha_fin = date(año, mes + 1, 1)

        with self.get_session() as session:
            ingresos = session.scalar(
                select(func.sum(Transaccion.cantidad))
                .where(Transaccion.tipo == TipoTransaccion.INGRESO)
                .where(Transaccion.fecha >= fecha_inicio)
                .where(Transaccion.fecha < fecha_fin)
            ) or Decimal("0")

            gastos = session.scalar(
                select(func.sum(Transaccion.cantidad))
                .where(Transaccion.tipo == TipoTransaccion.GASTO)
                .where(Transaccion.fecha >= fecha_inicio)
                .where(Transaccion.fecha < fecha_fin)
            ) or Decimal("0")

            return {
                "mes": mes,
                "año": año,
                "ingresos": ingresos,
                "gastos": gastos,
                "balance": ingresos - gastos,
            }

    # ==================== BUDGETS ====================

    def create_budget(
        self,
        nombre: str,
        categoria: str,
        limite: Decimal,
        periodo: PeriodoPresupuesto = PeriodoPresupuesto.MENSUAL,
        color: str = "#8B5CF6",
        icono: str = "fa-wallet",
        notas: str | None = None,
    ) -> Presupuesto:
        """Creates a new budget."""
        with self.get_session() as session:
            presupuesto = Presupuesto(
                nombre=nombre,
                categoria=categoria,
                limite=limite,
                periodo=periodo,
                color=color,
                icono=icono,
                notas=notas,
            )
            session.add(presupuesto)
            session.commit()
            session.refresh(presupuesto)
            return presupuesto

    def list_budgets(self, solo_activos: bool = True) -> list[Presupuesto]:
        """Lists all budgets."""
        with self.get_session() as session:
            query = select(Presupuesto)
            if solo_activos:
                query = query.where(Presupuesto.activo)
            return list(session.scalars(query).all())

    def get_budget(self, presupuesto_id: int) -> Presupuesto | None:
        """Gets a budget by ID."""
        with self.get_session() as session:
            return session.get(Presupuesto, presupuesto_id)

    def update_budget(
        self,
        presupuesto_id: int,
        nombre: str | None = None,
        limite: Decimal | None = None,
        color: str | None = None,
        icono: str | None = None,
        notas: str | None = None,
    ) -> Presupuesto | None:
        """Updates an existing budget."""
        with self.get_session() as session:
            presupuesto = session.get(Presupuesto, presupuesto_id)
            if not presupuesto:
                return None
            if nombre is not None:
                presupuesto.nombre = nombre
            if limite is not None:
                presupuesto.limite = limite
            if color is not None:
                presupuesto.color = color
            if icono is not None:
                presupuesto.icono = icono
            if notas is not None:
                presupuesto.notas = notas
            session.commit()
            session.refresh(presupuesto)
            return presupuesto

    def delete_budget(self, presupuesto_id: int) -> bool:
        """Deletes a budget."""
        with self.get_session() as session:
            presupuesto = session.get(Presupuesto, presupuesto_id)
            if presupuesto:
                session.delete(presupuesto)
                session.commit()
                return True
            return False

    def calculate_budget_spending(
        self, categoria: str, mes: int | None = None, año: int | None = None
    ) -> Decimal:
        """Calculates spending for a category in the selected period."""
        hoy = date.today()
        mes = mes or hoy.month
        año = año or hoy.year

        fecha_inicio = date(año, mes, 1)
        if mes == 12:
            fecha_fin = date(año + 1, 1, 1)
        else:
            fecha_fin = date(año, mes + 1, 1)

        with self.get_session() as session:
            gasto = session.scalar(
                select(func.sum(Transaccion.cantidad))
                .where(Transaccion.tipo == TipoTransaccion.GASTO)
                .where(Transaccion.categoria == categoria)
                .where(Transaccion.fecha >= fecha_inicio)
                .where(Transaccion.fecha < fecha_fin)
            ) or Decimal("0")
            return gasto

    def get_budget_status(self) -> list[dict]:
        """Gets the status of all active budgets."""
        presupuestos = self.list_budgets(solo_activos=True)
        resultado = []

        for p in presupuestos:
            gasto = self.calculate_budget_spending(p.categoria)
            limite = p.limite_mensual()
            porcentaje = (gasto / limite * 100) if limite > 0 else Decimal("0")
            disponible = limite - gasto

            resultado.append(
                {
                    "id": p.id,
                    "nombre": p.nombre,
                    "categoria": p.categoria,
                    "limite": float(limite),
                    "gastado": float(gasto),
                    "disponible": float(disponible),
                    "porcentaje": float(porcentaje),
                    "color": p.color,
                    "icono": p.icono,
                    "excedido": gasto > limite,
                }
            )

        return resultado

    # ==================== INSIGHTS ====================

    def get_insights(self) -> dict:
        """Generates financial insights and analytics."""
        today = date.today()

        # Gather data for the last 6 months.
        monthly_expenses = []
        monthly_income = []
        for i in range(6):
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 12
                year -= 1
            summary = self.monthly_summary(month, year)
            monthly_expenses.append(float(summary["gastos"]))
            monthly_income.append(float(summary["ingresos"]))

        # Calculate averages.
        average_expenses = sum(monthly_expenses) / len(monthly_expenses) if monthly_expenses else 0
        average_income = sum(monthly_income) / len(monthly_income) if monthly_income else 0

        # Current month summary.
        current_summary = self.monthly_summary(today.month, today.year)

        # Spending by category (current month).
        with self.get_session() as session:
            start_date = date(today.year, today.month, 1)
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1)
            else:
                end_date = date(today.year, today.month + 1, 1)

            expenses_by_category = session.execute(
                select(Transaccion.categoria, func.sum(Transaccion.cantidad).label("total"))
                .where(Transaccion.tipo == TipoTransaccion.GASTO)
                .where(Transaccion.fecha >= start_date)
                .where(Transaccion.fecha < end_date)
                .group_by(Transaccion.categoria)
                .order_by(func.sum(Transaccion.cantidad).desc())
            ).all()

        # Top spending category.
        top_category = expenses_by_category[0] if expenses_by_category else None

        # Savings rate.
        savings_rate = 0
        if current_summary["ingresos"] > 0:
            savings_rate = float(
                (current_summary["ingresos"] - current_summary["gastos"])
                / current_summary["ingresos"]
                * 100
            )

        # Comparison with previous month.
        previous_month = today.month - 1
        previous_year = today.year
        if previous_month <= 0:
            previous_month = 12
            previous_year -= 1
        previous_summary = self.monthly_summary(previous_month, previous_year)

        expenses_variation = 0
        if previous_summary["gastos"] > 0:
            expenses_variation = float(
                (current_summary["gastos"] - previous_summary["gastos"])
                / previous_summary["gastos"]
                * 100
            )

        # Remaining days in current month.
        if today.month == 12:
            last_day = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(today.year, today.month + 1, 1) - timedelta(days=1)
        remaining_days = (last_day - today).days

        # Average daily spending.
        elapsed_days = today.day
        daily_spending = float(current_summary["gastos"]) / elapsed_days if elapsed_days > 0 else 0

        # Monthly spending projection.
        projected_spending = daily_spending * last_day.day

        # Budgets at risk.
        budgets = self.get_budget_status()
        at_risk_budgets = [p for p in budgets if p["porcentaje"] >= 80]
        exceeded_budgets = [p for p in budgets if p["excedido"]]

        return {
            "resumen_mes": {
                "ingresos": float(current_summary["ingresos"]),
                "gastos": float(current_summary["gastos"]),
                "balance": float(current_summary["balance"]),
            },
            "promedios": {
                "gastos_mensual": average_expenses,
                "ingresos_mensual": average_income,
                "gasto_diario": daily_spending,
            },
            "comparacion_mes_anterior": {
                "variacion_gastos_porcentaje": expenses_variation,
                "gastos_mes_anterior": float(previous_summary["gastos"]),
            },
            "proyecciones": {
                "gasto_proyectado": projected_spending,
                "dias_restantes": remaining_days,
            },
            "tasa_ahorro": savings_rate,
            "categoria_mayor_gasto": {
                "nombre": top_category[0] if top_category else None,
                "total": float(top_category[1]) if top_category else 0,
            },
            "gastos_por_categoria": [
                {"categoria": c[0], "total": float(c[1])} for c in expenses_by_category
            ],
            "presupuestos": {
                "total": len(budgets),
                "en_riesgo": len(at_risk_budgets),
                "excedidos": len(exceeded_budgets),
                "detalle_riesgo": at_risk_budgets,
            },
            "alertas": self._generar_alertas(
                current_summary,
                expenses_variation,
                savings_rate,
                exceeded_budgets,
                at_risk_budgets,
            ),
        }

    def _generar_alertas(
        self,
        resumen: dict,
        variacion_gastos: float,
        tasa_ahorro: float,
        excedidos: list,
        en_riesgo: list,
    ) -> list[dict]:
        """Genera alertas basadas en el análisis financiero."""
        alertas = []

        # Alerta de gastos excesivos
        if variacion_gastos > 20:
            alertas.append(
                {
                    "tipo": "warning",
                    "icono": "fa-arrow-trend-up",
                    "titulo": "Aumento de gastos",
                    "mensaje": f"Tus gastos han aumentado un {variacion_gastos:.1f}% respecto al mes anterior.",
                }
            )

        # Alerta de balance negativo
        if resumen["balance"] < 0:
            alertas.append(
                {
                    "tipo": "danger",
                    "icono": "fa-triangle-exclamation",
                    "titulo": "Balance negativo",
                    "mensaje": "Estás gastando más de lo que ingresas este mes.",
                }
            )

        # Alerta de ahorro bajo
        if 0 <= tasa_ahorro < 10:
            alertas.append(
                {
                    "tipo": "warning",
                    "icono": "fa-piggy-bank",
                    "titulo": "Ahorro bajo",
                    "mensaje": f"Tu tasa de ahorro es del {tasa_ahorro:.1f}%. Considera reducir gastos.",
                }
            )

        # Alerta de buen ahorro
        if tasa_ahorro >= 20:
            alertas.append(
                {
                    "tipo": "success",
                    "icono": "fa-medal",
                    "titulo": "¡Excelente ahorro!",
                    "mensaje": f"Estás ahorrando el {tasa_ahorro:.1f}% de tus ingresos. ¡Sigue así!",
                }
            )

        # Alertas de presupuestos excedidos
        for p in excedidos:
            alertas.append(
                {
                    "tipo": "danger",
                    "icono": "fa-ban",
                    "titulo": f"Presupuesto excedido: {p['nombre']}",
                    "mensaje": f"Has superado tu límite de {p['limite']:.2f}€ en {p['categoria']}.",
                }
            )

        # Alertas de presupuestos en riesgo
        for p in en_riesgo:
            if not p["excedido"]:
                alertas.append(
                    {
                        "tipo": "warning",
                        "icono": "fa-clock",
                        "titulo": f"Presupuesto al límite: {p['nombre']}",
                        "mensaje": f"Has usado el {p['porcentaje']:.0f}% de tu presupuesto.",
                    }
                )

        return alertas


# English method aliases for backward-compatible translation
GestorDB.create_bank = GestorDB.create_bank
GestorDB.list_banks = GestorDB.list_banks
GestorDB.get_bank = GestorDB.get_bank
GestorDB.calculate_bank_balance = GestorDB.calculate_bank_balance

GestorDB.create_transaction = GestorDB.create_transaction
GestorDB.list_transactions = GestorDB.list_transactions
GestorDB.delete_transaction = GestorDB.delete_transaction

GestorDB.create_subscription = GestorDB.create_subscription
GestorDB.list_subscriptions = GestorDB.list_subscriptions
GestorDB.cancel_subscription = GestorDB.cancel_subscription
GestorDB.calculate_monthly_subscription_expense = GestorDB.calculate_monthly_subscription_cost

GestorDB.create_net_worth_item = GestorDB.create_net_worth_item
GestorDB.list_net_worth_items = GestorDB.list_net_worth_items
GestorDB.calculate_net_worth = GestorDB.calculate_net_worth

GestorDB.monthly_summary = GestorDB.monthly_summary
GestorDB.create_budget = GestorDB.create_budget
GestorDB.list_budgets = GestorDB.list_budgets
GestorDB.get_budget = GestorDB.get_budget
GestorDB.update_budget = GestorDB.update_budget
GestorDB.delete_budget = GestorDB.delete_budget
GestorDB.get_budget_status = GestorDB.get_budget_status
GestorDB.get_insights = GestorDB.get_insights


# English compatibility alias
DatabaseManager = GestorDB
